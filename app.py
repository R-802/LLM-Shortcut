"""
Headless hotkey service: Ctrl+B copies selection, queries LLM via meta-router, copies answer.

Uses Windows RegisterHotKey (no low-level keyboard hook).
Run at logon (hidden): pythonw.exe app.py
"""

from router.hotkey import load_hotkey_from_env
from router.images import (
    ImageAttachment,
    attachment_from_clipboard,
    collect_from_dirs,
)
from router.documents import read_pdf
from router.meta_router import complete
from router.rag import build_context_block, is_enabled as rag_enabled
from router.sanitize import sanitize_answer
import ctypes
import os
import sys
import time
import threading
import logging
from ctypes import wintypes
from pathlib import Path

import pyperclip
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

load_dotenv(APP_DIR / ".env")

LOG_PATH = APP_DIR / "app.log"
FOLDER_PATH = APP_DIR
CONTEXT_PATH = APP_DIR / "context"
NTFY_MAX_BODY = 3500
APP_BUILD = "2026-05-23-meta-router"
HOTKEY_ID = 1
MUTEX_NAME = "Global\\DataHotkeyExamHelper"
ERROR_ALREADY_EXISTS = 183

WM_HOTKEY = 0x0312
MOD_NOREPEAT = 0x4000

user32 = ctypes.windll.user32

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)


def _env_bool(name: str, default: bool = True) -> bool:
    raw = os.environ.get(name, "").strip().lower()
    if raw in ("1", "true", "yes", "on", "y"):
        return True
    if raw in ("0", "false", "no", "off", "n"):
        return False
    return default


NTFY_ENABLED = _env_bool("NTFY_ENABLED", default=True)
NTFY_TOPIC = os.environ.get("NTFY_TOPIC", "").strip()
if NTFY_ENABLED and not NTFY_TOPIC:
    logging.error(
        "NTFY_ENABLED is true but NTFY_TOPIC is empty. Set NTFY_TOPIC or set NTFY_ENABLED=false in %s",
        APP_DIR / ".env",
    )
    sys.exit(1)
if not NTFY_ENABLED:
    logging.info("ntfy notifications disabled (NTFY_ENABLED=false).")


_base_instruction = (
    "Respond in the style of a student writing a high-scoring exam answer.\n"
    "Use clear, direct language and standard academic terminology appropriate to the subject.\n"
    "Structure answers logically (e.g., brief paragraphs or bullet points where appropriate) without unnecessary elaboration.\n"
    "Avoid repetition, filler content, and background explanation unless it directly contributes to marks.\n"
    "Be very concise but complete: ensure all essential steps, definitions, or reasoning required for are included.\n"
    "Do not include commentary, teaching explanations, or meta-explanations.\n"
    "Do not include elaborations, examples, or non-essential details.\n"
    "Output plain text only with no formatting at all beyond basic line breaks and simple lists.\n"
    "Do not use markdown formatting.\n"
    "Do not use asterisks for headings or emphasis (use * only for multiplication in formulas).\n"
    "Do not use bold or italic formatting.\n"
    "Do not use quotes.\n"
    "Do not use code blocks.\n"
    "Do not use tables.\n"
    "Do not include images in your reply (text only).\n"
    "If images are attached, use them to answer the question.\n"
    "Do not use links.\n"
    "Do not use footnotes.\n"
    "Do not use endnotes.\n"
    "For maths use only ASCII operators: +, -, *, /, and powers as x^n (e.g. r^2, v^3).\n"
    "Do not use Unicode maths symbols (no pi, rho, multiply sign, divide sign, square root sign, etc.).\n"
    "Spell Greek letters in words (pi, rho, eta) - never use Greek symbol characters.\n"
    "Do not use LaTeX, markdown, or special formatting.\n"
    "Use New Zealand styled English.\n"
)

_busy = threading.Lock()


def _ensure_single_instance() -> None:
    kernel32 = ctypes.windll.kernel32
    kernel32.CreateMutexW(None, True, MUTEX_NAME)
    if kernel32.GetLastError() == ERROR_ALREADY_EXISTS:
        logging.error(
            "Another instance is already running. End pythonw.exe in Task Manager, then restart."
        )
        sys.exit(0)


def _ntfy_send(title: str, message: str, priority: str) -> None:
    if not NTFY_ENABLED:
        return
    url = f"https://ntfy.sh/{NTFY_TOPIC}"
    body = message if len(
        message) <= NTFY_MAX_BODY else message[:NTFY_MAX_BODY] + "\n..."
    headers = {"Title": title[:250], "Priority": priority, "Tags": "bell"}
    try:
        resp = requests.post(
            url,
            data=body.encode("utf-8"),
            headers=headers,
            timeout=15,
        )
        resp.raise_for_status()
        logging.info("ntfy sent: %s (HTTP %s)", title, resp.status_code)
    except requests.RequestException as exc:
        logging.error("ntfy failed (%s): %s", title, exc)


def ntfy_notify(title: str, message: str, priority: str = "3") -> None:
    """Push notification via ntfy.sh (runs in background thread)."""
    if not NTFY_ENABLED:
        return
    threading.Thread(
        target=_ntfy_send,
        args=(title, message, priority),
        daemon=True,
    ).start()


def _load_files_from_dir(folder: Path, file_data: str) -> str:
    skip_names = {
        "app.py", "app.log", ".env", "requirements.txt",
        "ROUTER_TEST.md", ".env.example",
    }
    skip_suffixes = {".pyc", ".yaml", ".md", ".bat", ".ps1", ".txt"}
    if not folder.is_dir():
        return file_data

    for file_path in folder.iterdir():
        if not file_path.is_file() or file_path.name in skip_names:
            continue
        if file_path.suffix in skip_suffixes or file_path.name.startswith("."):
            continue

        filename = file_path.name
        if filename.endswith(".pdf"):
            content = read_pdf(file_path)
            file_data += f"\n\nFile ({filename}):\n{content}"

        elif filename.endswith(".xlsx"):
            try:
                wb = load_workbook(file_path, data_only=True)
                text = ""
                for sheet in wb.worksheets:
                    text += f"\nSheet: {sheet.title}\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [
                            str(cell) if cell is not None else "" for cell in row
                        ]
                        text += " | ".join(row_text) + "\n"
                file_data += f"\n\nFile ({filename}):\n{text}"
            except Exception:
                file_data += f"\n\nFile ({filename}) could not be read."

        else:
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    file_data += f"\n\nFile ({filename}):\n" + f.read()
            except Exception:
                file_data += f"\n\nFile ({filename}) could not be read."

    return file_data


def load_context_files() -> str:
    """Load full file text (used when RAG is off or returns nothing)."""
    file_data = ""
    file_data = _load_files_from_dir(FOLDER_PATH, file_data)
    if CONTEXT_PATH.is_dir():
        file_data = _load_files_from_dir(CONTEXT_PATH, file_data)
    return file_data


def load_prompt_context(question: str) -> str:
    """RAG retrieval from context/ when enabled; otherwise dump context files."""
    if rag_enabled() and CONTEXT_PATH.is_dir():
        rag_text = build_context_block(question, CONTEXT_PATH, APP_DIR)
        if rag_text.strip():
            return rag_text
        logging.info(
            "RAG found no chunks; falling back to full context files.")
    return load_context_files()


def load_context_images() -> list[ImageAttachment]:
    return collect_from_dirs(FOLDER_PATH, CONTEXT_PATH)


def capture_from_clipboard() -> tuple[str, list[ImageAttachment]]:
    """
    Read whatever is on the clipboard when the user presses the hotkey.

    Does not send Ctrl+C or any other keys. Copy your selection (Ctrl+C) and/or
    take a screenshot (Win+Shift+S) first, then press your configured shortcut.
    """
    text = pyperclip.paste().strip()

    images: list[ImageAttachment] = []
    clip_image = attachment_from_clipboard(label="clipboard_image")
    if clip_image:
        images.append(clip_image)
        logging.info("Captured clipboard image (%s).", clip_image.label)

    if text:
        logging.info("Captured %d chars of clipboard text.", len(text))
    return text, images


def ask_llm(question: str, images: list[ImageAttachment] | None = None) -> tuple[str, str]:
    """Return (answer, provider_id) via meta-router."""
    file_data = load_prompt_context(question)
    user = f"Question:\n{question}\n\nContext:\n{file_data}"

    all_images: list[ImageAttachment] = []
    all_images.extend(load_context_images())
    if images:
        for img in images:
            if not any(i.label == img.label for i in all_images):
                all_images.append(img)
    if len(all_images) > 6:
        all_images = all_images[:6]

    answer, meta = complete(
        system=_base_instruction,
        user_prompt=user,
        images=all_images or None,
    )
    answer = sanitize_answer(answer.strip())
    logging.info(
        "Routed tier=%s provider=%s model=%s attempts=%d escalated=%s "
        "instruction_chars=%d images=%d rag=%s",
        meta.tier,
        meta.provider_id,
        meta.litellm_model,
        meta.attempts,
        meta.escalated,
        len(_base_instruction),
        len(all_images),
        rag_enabled(),
    )
    return answer, meta.provider_id


def _handle_hotkey() -> None:
    if not _busy.acquire(blocking=False):
        logging.info("Hotkey ignored; previous request still running.")
        ntfy_notify("Exam helper",
                    "Still working on the previous question.", priority="3")
        return

    try:
        question, clip_images = capture_from_clipboard()
        context_images = load_context_images()
        has_images = bool(clip_images or context_images)

        if not question and not has_images:
            pyperclip.copy("No text or image selected.")
            logging.warning("Empty selection.")
            ntfy_notify("Exam helper",
                        "No text or image selected.", priority="3")
            return

        if not question and has_images:
            question = "Answer the exam question shown in the image(s)."

        preview = question[:120] + ("…" if len(question) > 120 else "")
        ntfy_notify("Exam helper", f"Working on:\n{preview}", priority="1")

        logging.info(
            "Sending %d chars, %d image(s) to meta-router.",
            len(question),
            len(clip_images) + len(context_images),
        )
        answer, provider_id = ask_llm(question, images=clip_images)
        pyperclip.copy(answer or "No response from model.")
        logging.info("Answer copied (%d chars) via %s.",
                     len(answer), provider_id)

        footer = f"\n\n— via {provider_id}"
        body = (answer or "No response from model. Paste from clipboard on your PC.")
        if len(body) + len(footer) <= NTFY_MAX_BODY:
            body += footer
        ntfy_notify("Answer ready", body, priority="max")
    except Exception as exc:
        pyperclip.copy(f"Error: {exc}")
        logging.exception("Hotkey handler failed.")
        ntfy_notify("Exam helper failed", str(exc), priority="5")
    finally:
        _busy.release()


def _on_hotkey() -> None:
    threading.Thread(target=_handle_hotkey, daemon=True).start()


def main() -> None:
    _ensure_single_instance()
    logging.info("Build %s | python=%s | cwd=%s",
                 APP_BUILD, sys.executable, os.getcwd())

    try:
        hotkey_mod, hotkey_vk, hotkey_display = load_hotkey_from_env()
    except ValueError as exc:
        logging.error("%s", exc)
        sys.exit(1)

    modifiers = hotkey_mod | MOD_NOREPEAT
    if not user32.RegisterHotKey(None, HOTKEY_ID, modifiers, hotkey_vk):
        logging.error(
            "RegisterHotKey(%s) failed. Another program may already own this shortcut.",
            hotkey_display,
        )
        sys.exit(1)

    logging.info(
        "Hotkey service started (%s via RegisterHotKey).", hotkey_display)
    ntfy_notify(
        "Exam helper started",
        f"Build {APP_BUILD} is running. {hotkey_display} is active.",
        "4",
    )

    msg = wintypes.MSG()
    try:
        while user32.GetMessageW(ctypes.byref(msg), None, 0, 0) > 0:
            if msg.message == WM_HOTKEY and msg.wParam == HOTKEY_ID:
                _on_hotkey()
            user32.TranslateMessage(ctypes.byref(msg))
            user32.DispatchMessageW(ctypes.byref(msg))
    finally:
        user32.UnregisterHotKey(None, HOTKEY_ID)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
