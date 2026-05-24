"""Extract plain text from study materials for indexing and context."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

# Files never indexed as RAG sources
SKIP_NAMES = {
    "app.py", "app.log", ".env", "requirements.txt",
    "ROUTER_TEST.md", ".env.example",
}

SKIP_SUFFIXES = {".pyc", ".yaml", ".bat", ".ps1", ".log"}

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}

INDEXABLE_SUFFIXES = {
    ".pdf", ".xlsx", ".xls", ".txt", ".md", ".csv", ".json", ".py",
    ".html", ".htm", ".docx",
}

_MAX_TABLE_ROWS = 120
_TIMESERIES_SAMPLE_ROWS = 8


def read_pdf(file_path: Path) -> str:
    text = ""
    try:
        reader = PdfReader(str(file_path))
        for page in reader.pages:
            text += page.extract_text() or ""
    except Exception:
        text = "PDF could not be read."
    return text


def _normalize_cell(cell: object) -> str:
    if cell is None:
        return ""
    if isinstance(cell, bool):
        return "TRUE" if cell else "FALSE"
    if isinstance(cell, int):
        return str(cell)
    if isinstance(cell, float):
        if cell == int(cell):
            return str(int(cell))
        return str(cell)
    if isinstance(cell, (datetime, date)):
        if isinstance(cell, datetime):
            return cell.strftime("%Y-%m-%d %H:%M")
        return cell.isoformat()

    text = str(cell).strip()
    if not text:
        return ""

    # Common degree-symbol mojibake from Excel (e.g. b"C" or replacement char + C)
    if re.fullmatch(r"[\u00b0\u00ba\ufffd]?\s*C", text, re.IGNORECASE):
        return "deg C"
    text = text.replace("\ufffd", "")
    text = text.replace("\u00b0C", "deg C").replace("\u00baC", "deg C")
    text = text.replace("ºC", "deg C")
    return text


def _rows_to_csv(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().strip()


def _pad_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []
    width = max(len(r) for r in rows)
    return [r + [""] * (width - len(r)) for r in rows]


def _used_column_indices(rows: list[list[str]]) -> list[int]:
    padded = _pad_rows(rows)
    if not padded:
        return []
    width = len(padded[0])
    return [
        ci
        for ci in range(width)
        if any(padded[ri][ci].strip() for ri in range(len(padded)))
    ]


def _column_regions(used_cols: list[int]) -> list[tuple[int, int]]:
    if not used_cols:
        return []
    regions: list[tuple[int, int]] = []
    start = used_cols[0]
    prev = used_cols[0]
    for col in used_cols[1:]:
        if col - prev > 1:
            regions.append((start, prev))
            start = col
        prev = col
    regions.append((start, prev))
    return regions


def _slice_region(rows: list[list[str]], col_start: int, col_end: int) -> list[list[str]]:
    sliced: list[list[str]] = []
    for row in rows:
        cells = [
            row[i].strip() if i < len(row) else ""
            for i in range(col_start, col_end + 1)
        ]
        if any(cells):
            sliced.append(cells)
    return sliced


def _row_looks_like_timestamp(cell: str) -> bool:
    if not cell:
        return False
    if re.match(r"20\d{2}[-/]\d", cell):
        return True
    return ":" in cell and len(cell) >= 10 and any(ch.isdigit() for ch in cell)


def _looks_like_timeseries(rows: list[list[str]]) -> bool:
    if len(rows) < 30:
        return False
    data_rows = rows[1:] if len(rows) > 1 else rows
    ts_rows = sum(1 for r in data_rows if r and _row_looks_like_timestamp(r[0]))
    return ts_rows >= max(20, int(len(rows) * 0.35))


def _region_label(rows: list[list[str]], col_start: int, col_end: int) -> str:
    preview = " ".join(
        c for r in rows[:3] for c in r if c
    ).lower()
    if _looks_like_timeseries(rows):
        return "timeseries data"
    if (col_end - col_start) <= 2 and ("=" in preview or "assumption" in preview):
        return "inputs and assumptions"
    if "coefficient" in preview or "calc" in preview:
        return "calculations"
    if any(w in preview for w in ("conclusion", "nah", "recommend")):
        return "conclusions"
    return "table"


def _format_kv_region(rows: list[list[str]]) -> list[list[str]]:
    """Two-column field/value layout common in tutorial spreadsheets."""
    out: list[list[str]] = [["field", "value"]]
    for row in rows:
        cells = [c for c in row if c]
        if not cells:
            continue
        if len(cells) == 1:
            out.append([cells[0], ""])
        elif len(cells) == 2:
            out.append([cells[0], cells[1]])
        else:
            out.append([cells[0], " ".join(cells[1:])])
    return out


def _format_table_region(rows: list[list[str]]) -> list[list[str]]:
    """Keep a header row when present; cap long tables."""
    if not rows:
        return []
    non_empty = [r for r in rows if any(c for c in r)]
    if not non_empty:
        return []

    header = non_empty[0]
    body = non_empty[1:]
    if len(non_empty) > _MAX_TABLE_ROWS:
        body = body[:_MAX_TABLE_ROWS]
        return [header, *body, [f"... ({len(non_empty) - 1 - _MAX_TABLE_ROWS} more rows omitted)"]]
    return non_empty


def _format_timeseries_region(rows: list[list[str]]) -> str:
    header = rows[0]
    body = rows[1:]
    sample = body[:_TIMESERIES_SAMPLE_ROWS]
    lines = [
        f"timeseries ({len(body)} rows; sample only)",
        _rows_to_csv([header, *sample]),
    ]
    if len(body) > _TIMESERIES_SAMPLE_ROWS:
        lines.append(f"... ({len(body) - _TIMESERIES_SAMPLE_ROWS} hourly rows omitted)")
    return "\n".join(lines)


def _format_region(rows: list[list[str]], label: str) -> str:
    if _looks_like_timeseries(rows):
        return f"--- {label} ---\n{_format_timeseries_region(rows)}"

    if len(rows[0]) <= 3:
        table = _format_kv_region(rows)
    else:
        table = _format_table_region(rows)

    if not table:
        return ""
    return f"--- {label} ---\n{_rows_to_csv(table)}"


def _extract_sheet(sheet) -> str:
    raw_rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        raw_rows.append([_normalize_cell(c) for c in row])

    raw_rows = [r for r in raw_rows if any(c for c in r)]
    if not raw_rows:
        return ""

    used = _used_column_indices(raw_rows)
    if not used:
        return ""

    parts: list[str] = [f"=== Sheet: {sheet.title} ==="]
    regions = _column_regions(used)

    for col_start, col_end in regions:
        region_rows = _slice_region(raw_rows, col_start, col_end)
        if not region_rows:
            continue
        label = _region_label(region_rows, col_start, col_end)
        block = _format_region(region_rows, label)
        if block:
            parts.append(block)

    return "\n\n".join(parts)


def _extract_xlsx(file_path: Path) -> str:
    """Extract xlsx as labelled CSV blocks (separate regions per sheet)."""
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception:
        return ""

    parts: list[str] = []
    for sheet in wb.worksheets:
        block = _extract_sheet(sheet)
        if block:
            parts.append(block)

    return "\n\n".join(parts)


def extract_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(file_path)
    if suffix in {".xlsx", ".xls"}:
        return _extract_xlsx(file_path)
    try:
        return file_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def iter_source_files(source_dir: Path) -> list[Path]:
    if not source_dir.is_dir():
        return []
    files: list[Path] = []
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.name in SKIP_NAMES or path.name.startswith("."):
            continue
        if path.suffix.lower() in SKIP_SUFFIXES | IMAGE_SUFFIXES:
            continue
        if path.suffix.lower() not in INDEXABLE_SUFFIXES:
            continue
        files.append(path)
    return files
